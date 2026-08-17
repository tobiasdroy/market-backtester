"""Bank of England "A Millennium of Macroeconomic Data for the UK" —
free, centuries of UK data (v3.1, runs to ~2017). Used as the long-run
proxy for UK inflation (pre-1914 annual, 1914-1987 monthly), long gilt
yields (pre-1960), and GBP Bank Rate (pre-2016).

Source: https://www.bankofengland.co.uk/statistics/research-datasets
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent.parent))
from common import download_cached, month_start  # noqa: E402

BOE_URL = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/"
    "research-datasets/a-millennium-of-macroeconomic-data-for-the-uk.xlsx"
)

MONTH_NUM = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12,
}

# Column indices (0-based) within the "M1. Mthly headline series" sheet.
M1_COL_CPI = 7  # Spliced monthly CPI, 1914-2015, 2015=100
M1_COL_BANK_RATE = 11  # Bank Rate 1694-2015, % pa
M1_COL_CONSOLS_YIELD = 20  # Long-term consols yield 1753-2015, % pa
M1_COL_FX_USD_PER_GBP = 24  # $/£ exchange rate

# Column index within "A47. Wages and prices" (annual sheet).
A47_COL_CPI = 3  # Consumer Price Index (CPI) - preferred measure, 2015=100


def _load_workbook():
    path = download_cached(BOE_URL, "boe_millennium.xlsx")
    return openpyxl.load_workbook(str(path), read_only=True, data_only=True)


def _m1_monthly_column(ws, col: int) -> pd.Series:
    out: dict[pd.Timestamp, float] = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        year = row[0]
        month_name = row[1]
        if not isinstance(year, int) or month_name not in MONTH_NUM:
            continue
        val = row[col]
        if val is None or not isinstance(val, (int, float)):
            continue
        out[month_start(year, MONTH_NUM[month_name])] = float(val)
    return pd.Series(out).sort_index()


def fetch_boe_m1_cpi() -> pd.Series:
    wb = _load_workbook()
    return _m1_monthly_column(wb["M1. Mthly headline series"], M1_COL_CPI)


def fetch_boe_m1_bank_rate_pct() -> pd.Series:
    wb = _load_workbook()
    return _m1_monthly_column(wb["M1. Mthly headline series"], M1_COL_BANK_RATE)


def fetch_boe_m1_consols_yield_pct() -> pd.Series:
    wb = _load_workbook()
    return _m1_monthly_column(wb["M1. Mthly headline series"], M1_COL_CONSOLS_YIELD)


def fetch_boe_m1_fx_usd_per_gbp() -> pd.Series:
    wb = _load_workbook()
    return _m1_monthly_column(wb["M1. Mthly headline series"], M1_COL_FX_USD_PER_GBP)


def fetch_boe_a47_annual_cpi() -> pd.Series:
    """Annual UK CPI index level (2015=100), indexed by year int."""
    wb = _load_workbook()
    ws = wb["A47. Wages and prices"]
    out: dict[int, float] = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        year = row[0]
        if not isinstance(year, int):
            continue
        val = row[A47_COL_CPI]
        if val is None or not isinstance(val, (int, float)):
            continue
        out[year] = float(val)
    return pd.Series(out).sort_index()
